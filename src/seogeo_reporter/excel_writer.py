from __future__ import annotations

import shutil
from pathlib import Path

from .models import MonthlyReportPayload


class ExcelWriterError(RuntimeError):
    pass


class SeoGeoExcelWriter:
    """Writes payload data into SEOGEO workbook template.

    Uses openpyxl when available. Keeps formulas/style by copying template first.
    """

    def __init__(self, template_path: str = "SEOGEO_리포트양식.xlsx") -> None:
        self.template_path = Path(template_path)

    def write(self, payload: MonthlyReportPayload, output_path: str) -> Path:
        if not self.template_path.exists():
            raise ExcelWriterError(f"Template not found: {self.template_path}")

        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as exc:
            raise ExcelWriterError(
                "openpyxl is required for Excel writing. Install openpyxl in runtime environment."
            ) from exc

        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(self.template_path, out)

        wb = load_workbook(out)

        self._write_overview(wb, payload)
        self._write_prompt_rows(wb, payload)
        self._write_keyword_rows(wb, payload)
        self._write_summary(wb, payload)

        wb.save(out)
        return out

    @staticmethod
    def _write_overview(wb, payload: MonthlyReportPayload) -> None:
        if "리포트개요" not in wb.sheetnames:
            return
        ws = wb["리포트개요"]
        ws["B3"] = payload.overview.get("client_name")
        ws["B4"] = payload.overview.get("site_url")
        ws["B5"] = payload.overview.get("month")

    @staticmethod
    def _write_prompt_rows(wb, payload: MonthlyReportPayload) -> None:
        if "GEO프롬프트테스트결과" not in wb.sheetnames:
            return
        ws = wb["GEO프롬프트테스트결과"]
        start = 5
        for i, row in enumerate(payload.prompt_rows):
            r = start + i
            ws[f"A{r}"] = row.get("prompt_id")
            ws[f"B{r}"] = row.get("engine")
            ws[f"C{r}"] = row.get("capture_method")
            ws[f"D{r}"] = row.get("prompt_text")
            ws[f"E{r}"] = row.get("tier")
            ws[f"F{r}"] = row.get("tier_score")
            ws[f"G{r}"] = row.get("screenshot_path")

    @staticmethod
    def _write_keyword_rows(wb, payload: MonthlyReportPayload) -> None:
        if "중위키워드추적" not in wb.sheetnames:
            return
        ws = wb["중위키워드추적"]
        start = 5
        for i, row in enumerate(payload.keyword_rows):
            r = start + i
            ws[f"A{r}"] = row.get("keyword")
            ws[f"B{r}"] = row.get("previous_impressions")
            ws[f"C{r}"] = row.get("current_impressions")
            ws[f"D{r}"] = row.get("previous_clicks")
            ws[f"E{r}"] = row.get("current_clicks")
            ws[f"F{r}"] = row.get("previous_ctr")
            ws[f"G{r}"] = row.get("current_ctr")
            ws[f"H{r}"] = row.get("previous_position")
            ws[f"I{r}"] = row.get("current_position")

    @staticmethod
    def _write_summary(wb, payload: MonthlyReportPayload) -> None:
        if "월간종합리포트" not in wb.sheetnames or not payload.summary_rows:
            return
        ws = wb["월간종합리포트"]
        first = payload.summary_rows[0]
        ws["B5"] = first.get("organic_sessions")
        ws["B6"] = first.get("referral_sessions")
        ws["B7"] = first.get("conversions")

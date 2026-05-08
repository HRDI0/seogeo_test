from __future__ import annotations

import csv
from pathlib import Path


class PromptInputError(ValueError):
    pass


def load_prompt_set(prompt_file: str | None, inline_prompts: list[str]) -> list[dict[str, str]]:
    prompts: list[str] = []
    if prompt_file:
        prompts.extend(_load_prompts_from_file(prompt_file))
    prompts.extend([p.strip() for p in inline_prompts if p.strip()])

    if not prompts:
        raise PromptInputError("at least one prompt is required via --prompt-file or --prompt")

    return [
        {
            "prompt_id": f"P{i:03d}",
            "prompt_type": "직접형",
            "prompt_text": text,
        }
        for i, text in enumerate(prompts, start=1)
    ]


def _load_prompts_from_file(path: str) -> list[str]:
    p = Path(path)
    if not p.exists():
        raise PromptInputError(f"prompt file not found: {path}")

    ext = p.suffix.lower()
    if ext == ".csv":
        return _load_from_csv(p)
    if ext in {".xlsx", ".xlsm"}:
        return _load_from_xlsx(p)
    raise PromptInputError("prompt file must be .csv or .xlsx/.xlsm")


def _load_from_csv(path: Path) -> list[str]:
    with path.open("r", encoding="utf-8-sig", newline="") as fp:
        reader = csv.DictReader(fp)
        return _rows_to_prompts(reader)


def _load_from_xlsx(path: Path) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise PromptInputError("openpyxl is required to load .xlsx prompt files") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return []
    keys = [str(v).strip().lower() if v else "" for v in header]
    payload = []
    for row in rows:
        payload.append({keys[i]: row[i] for i in range(min(len(keys), len(row)))})
    return _rows_to_prompts(payload)


def _rows_to_prompts(rows) -> list[str]:
    prompts: list[str] = []
    for row in rows:
        value = row.get("prompt_text") or row.get("prompt") or row.get("query")
        if value:
            prompts.append(str(value).strip())
    return [p for p in prompts if p]
